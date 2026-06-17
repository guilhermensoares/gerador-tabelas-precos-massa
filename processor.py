from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from functools import lru_cache
from io import BytesIO
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple
import re
import zipfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


@dataclass
class ProcessConfig:
    tabela_codigo: str = "007"
    tabela_codigo_013: str = "013"
    acrescimo_tabela_013: float = 0.12
    desconto_ml: float = 0.10
    markup_minimo: float = 1.20
    margem_critica: float = 0.35
    margem_atencao: float = 0.50
    tolerancia: float = 0.009
    preco_007_sem_cadastro: float = 0.01
    usar_ocorrencia_duplicada: str = "ultima"  # ultima | primeira


@dataclass
class ProcessResult:
    analise: pd.DataFrame
    alertas: pd.DataFrame
    log: pd.DataFrame
    resumo: pd.DataFrame
    comparativo: pd.DataFrame
    protheus: pd.DataFrame
    protheus_013: pd.DataFrame
    matched_table: pd.DataFrame
    resumo_dict: Dict[str, int | float | str]
    log_text: str


# -----------------------------
# Normalização e parsing
# -----------------------------

def normalize_sku(value) -> Optional[str]:
    """Normaliza SKU King para 5 dígitos, preservando zeros à esquerda."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none", "null"}:
        return None

    # Evita que 12345.0 vire 123450
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]

    # Para SKUs numéricos, mantém apenas dígitos e completa para 5 posições
    digits = re.sub(r"\D", "", text)
    if digits:
        return digits.zfill(5)[-5:] if len(digits) <= 5 else digits

    # Fallback: retorna texto original caso algum SKU não seja puramente numérico
    return text.upper()




def validate_sku_format(value) -> Tuple[Optional[str], str, str]:
    """Valida e normaliza SKU King.

    Retorna: (sku_normalizado, severidade, mensagem)
    Severidades: OK | ATENÇÃO | CRÍTICO
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, "CRÍTICO", "SKU vazio/ausente."

    raw = str(value).strip()
    if raw == "" or raw.lower() in {"nan", "none", "null"}:
        return None, "CRÍTICO", "SKU vazio/ausente."

    # Excel pode transformar 01234 em 1234 ou 1234.0. Este caso é aceito e normalizado.
    numeric_decimal = re.fullmatch(r"\d+\.0", raw)
    text_for_digits = raw[:-2] if numeric_decimal else raw
    digits = re.sub(r"\D", "", text_for_digits)

    if not digits:
        return None, "CRÍTICO", f"SKU sem dígitos válidos: {raw!r}."

    if len(digits) > 5:
        return None, "CRÍTICO", f"SKU com mais de 5 dígitos ({raw!r}). Não será enviado ao Protheus para evitar cadastro incorreto."

    sku = digits.zfill(5)

    # Aceita formatos numéricos puros e numéricos com .0 gerados pelo Excel.
    if re.fullmatch(r"\d{1,5}", text_for_digits) or numeric_decimal:
        return sku, "OK", ""

    return sku, "ATENÇÃO", f"SKU em formato não padrão ({raw!r}); normalizado para {sku}."

def parse_money(value) -> Optional[float]:
    """Converte valores no padrão BR/US para float. Retorna None quando inválido."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none", "null", "-"}:
        return None

    text = text.replace("R$", "").replace(" ", "")
    text = re.sub(r"[^0-9,\.\-]", "", text)
    if text in {"", "-", ".", ","}:
        return None

    # Se tiver vírgula e ponto, assume que o último separador indica decimal.
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(".", "").replace(",", ".")

    try:
        return float(text)
    except ValueError:
        return None


def money_dot(value: Optional[float]) -> str:
    if value is None or pd.isna(value):
        return ""
    return f"{float(value):.2f}"


def pct(value: Optional[float]) -> Optional[float]:
    if value is None or pd.isna(value):
        return None
    return float(value)


def norm_col_name(col) -> str:
    text = str(col).strip().lower()
    text = text.replace("ç", "c").replace("ã", "a").replace("á", "a").replace("à", "a")
    text = text.replace("â", "a").replace("é", "e").replace("ê", "e").replace("í", "i")
    text = text.replace("ó", "o").replace("ô", "o").replace("õ", "o").replace("ú", "u")
    text = re.sub(r"\s+", " ", text)
    return text


def detect_column(df: pd.DataFrame, candidates: Iterable[str], fallback_index: Optional[int] = None) -> str:
    norm_map = {norm_col_name(c): c for c in df.columns}
    cand_norms = [norm_col_name(c) for c in candidates]

    # match exato
    for cand in cand_norms:
        if cand in norm_map:
            return norm_map[cand]

    # match por contenção
    for cand in cand_norms:
        for norm, original in norm_map.items():
            if cand in norm or norm in cand:
                return original

    if fallback_index is not None and fallback_index < len(df.columns):
        return df.columns[fallback_index]

    raise ValueError(f"Não foi possível encontrar coluna: {list(candidates)}")


def read_excel_flexible(uploaded_file, preferred_sheet: Optional[str] = None) -> Tuple[pd.DataFrame, str]:
    """Lê Excel aceitando caminho, BytesIO ou arquivo do Streamlit."""
    xls = pd.ExcelFile(_excel_source(uploaded_file))
    if preferred_sheet and preferred_sheet in xls.sheet_names:
        sheet = preferred_sheet
    else:
        # Evita abas de resumo/log como fonte principal, quando possível
        candidates = [s for s in xls.sheet_names if not s.upper().startswith(("RESUMO", "LOG", "ALERT", "COMPAR"))]
        sheet = candidates[0] if candidates else xls.sheet_names[0]
    df = pd.read_excel(xls, sheet_name=sheet, dtype=object)
    df = df.dropna(how="all")
    return df, sheet




def _safe_seek(file_obj):
    """Volta o ponteiro do arquivo para o início quando possível.
    Necessário para arquivos enviados pelo Streamlit, que podem ser lidos mais de uma vez.
    """
    try:
        file_obj.seek(0)
    except Exception:
        pass


def _excel_source(file_obj):
    """Converte bytes em BytesIO e reposiciona streams já abertos."""
    if isinstance(file_obj, (bytes, bytearray)):
        return BytesIO(bytes(file_obj))
    _safe_seek(file_obj)
    return file_obj


def _file_bytes_or_none(file_obj):
    """Retorna bytes quando o arquivo permite cache seguro em memória."""
    if isinstance(file_obj, (bytes, bytearray)):
        return bytes(file_obj)
    try:
        return file_obj.getvalue()
    except Exception:
        return None


def _path_cache_key(file_obj):
    try:
        path = Path(file_obj)
        if path.exists():
            stat = path.stat()
            return str(path.resolve()), stat.st_mtime_ns, stat.st_size
    except Exception:
        return None
    return None


def _read_fast_tabela_007(tabela_007_file) -> Tuple[pd.DataFrame, list, str]:
    """Leitura rápida da tabela 007.

    A rotina oficial usa a aba "2-Produtos da tabela de preço", SKU na coluna G
    e preço na coluna L. Lemos apenas essas colunas com openpyxl em modo read_only,
    evitando carregar as 17 colunas e 35k linhas no pandas.
    """
    _safe_seek(tabela_007_file)
    wb = load_workbook(_excel_source(tabela_007_file), read_only=True, data_only=True)
    sheet = "2-Produtos da tabela de preço" if "2-Produtos da tabela de preço" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet]

    rows = []
    log = []
    seen = set()
    duplicated = set()

    # Coluna G = 7; Coluna L = 12.
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, min_col=7, max_col=12, values_only=True), start=2):
        raw_sku = values[0]
        raw_price = values[5]
        sku, sev, msg = validate_sku_format(raw_sku)
        if sku is None and (raw_sku is None or str(raw_sku).strip().lower() in {"", "nan", "none", "null"}):
            continue

        if sev != "OK":
            log.append({
                "Tipo": "FORMATO_SKU_TABELA_007",
                "Severidade": sev,
                "SKU": sku or "",
                "Mensagem": "Formato de SKU na tabela 007 exige revisão." if sev == "ATENÇÃO" else "Formato de SKU inválido na tabela 007.",
                "Detalhe": f"Linha Excel aproximada: {excel_row}; valor original: {raw_sku}; {msg}",
            })

        if sku:
            if sku in seen:
                duplicated.add(sku)
            seen.add(sku)
            rows.append({
                "SKU_NORMALIZADO": sku,
                "SKU_SEVERIDADE": sev,
                "SKU_MENSAGEM": msg,
                "SKU_ORIGINAL": str(raw_sku),
                "PRECO_TABELA_007": parse_money(raw_price),
            })

    try:
        wb.close()
    except Exception:
        pass

    if duplicated:
        log.append({
            "Tipo": "DUPLICIDADE_TABELA_007",
            "Severidade": "ATENÇÃO",
            "SKU": "",
            "Mensagem": f"Há {len(duplicated)} SKU(s) duplicado(s) na tabela 007.",
            "Detalhe": "O app usará a primeira ocorrência para cálculo. Revise se houver preços diferentes.",
        })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.drop_duplicates("SKU_NORMALIZADO", keep="first")
    return df, log, sheet


def _read_fast_lista_ml(lista_ml_file, config: ProcessConfig) -> Tuple[pd.DataFrame, list, str]:
    """Leitura rápida da Lista ML revisada.

    Estrutura oficial atual: A=SKU, B=Preço ML, C=Custo Médio.
    """
    _safe_seek(lista_ml_file)
    wb = load_workbook(_excel_source(lista_ml_file), read_only=True, data_only=True)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]

    rows = []
    log = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True), start=2):
        raw_sku, raw_ml, raw_custo = values
        sku, sev, msg = validate_sku_format(raw_sku)
        if sku is None and (raw_sku is None or str(raw_sku).strip().lower() in {"", "nan", "none", "null"}):
            continue
        if sev != "OK":
            log.append({
                "Tipo": "FORMATO_SKU_LISTA_ML",
                "Severidade": sev,
                "SKU": sku or "",
                "Mensagem": "Formato de SKU na Lista ML revisada exige revisão." if sev == "ATENÇÃO" else "Formato de SKU inválido na Lista ML revisada.",
                "Detalhe": f"Linha Excel aproximada: {excel_row}; valor original: {raw_sku}; {msg}",
            })
        if sku:
            rows.append({"SKU": sku, "PRECO_ML": parse_money(raw_ml), "CUSTO_MEDIO": parse_money(raw_custo)})

    try:
        wb.close()
    except Exception:
        pass

    df = pd.DataFrame(rows, columns=["SKU", "PRECO_ML", "CUSTO_MEDIO"])
    if df.empty:
        return df, log, sheet

    dupes = df[df.duplicated("SKU", keep=False)]
    if not dupes.empty:
        for sku, grp in dupes.groupby("SKU"):
            prices = sorted(set([round(x, 2) for x in grp["PRECO_ML"].dropna().tolist()]))
            costs = sorted(set([round(x, 2) for x in grp["CUSTO_MEDIO"].dropna().tolist()]))
            log.append({
                "Tipo": "DUPLICIDADE_LISTA_ML",
                "Severidade": "ATENÇÃO",
                "SKU": sku,
                "Mensagem": "SKU duplicado na Lista ML revisada.",
                "Detalhe": f"Preços ML encontrados: {prices}; custos encontrados: {costs}. Usada a {config.usar_ocorrencia_duplicada} ocorrência.",
            })
        keep = "last" if config.usar_ocorrencia_duplicada == "ultima" else "first"
        df = df.drop_duplicates("SKU", keep=keep)

    return df[["SKU", "PRECO_ML", "CUSTO_MEDIO"]], log, sheet


@lru_cache(maxsize=4)
def _cached_fast_tabela_007_from_bytes(file_bytes: bytes) -> Tuple[pd.DataFrame, list, str]:
    return _read_fast_tabela_007(BytesIO(file_bytes))


@lru_cache(maxsize=4)
def _cached_fast_tabela_007_from_path(path: str, mtime_ns: int, size: int) -> Tuple[pd.DataFrame, list, str]:
    return _read_fast_tabela_007(path)


def _read_fast_tabela_007_cached(tabela_007_file) -> Tuple[pd.DataFrame, list, str]:
    """Cacheia a leitura da tabela 007.

    A base oficial tem dezenas de milhares de linhas e normalmente é reaproveitada
    por vários processamentos. Cachear essa etapa deixa as próximas execuções bem
    mais rápidas quando a mesma tabela 007 é usada novamente.
    """
    file_bytes = _file_bytes_or_none(tabela_007_file)
    if file_bytes is not None:
        df, log, sheet = _cached_fast_tabela_007_from_bytes(file_bytes)
        return df.copy(), [dict(x) for x in log], sheet

    key = _path_cache_key(tabela_007_file)
    if key is not None:
        df, log, sheet = _cached_fast_tabela_007_from_path(*key)
        return df.copy(), [dict(x) for x in log], sheet

    return _read_fast_tabela_007(tabela_007_file)


# -----------------------------
# Preparação das bases
# -----------------------------

def prepare_tabela_007(tabela_007_file) -> Tuple[pd.DataFrame, pd.DataFrame, list, str]:
    """Prepara tabela 007.

    Usa leitura rápida por colunas fixas. Se a estrutura do arquivo mudar, cai no método
    flexível antigo com pandas para tentar localizar as colunas pelo cabeçalho.
    """
    try:
        calc, log, sheet = _read_fast_tabela_007_cached(tabela_007_file)
        return calc, pd.DataFrame(), log, sheet
    except Exception:
        _safe_seek(tabela_007_file)
        raw, sheet = read_excel_flexible(tabela_007_file, preferred_sheet="2-Produtos da tabela de preço")
        sku_col = detect_column(raw, ["Cod.Produto", "Cod Produto", "SKU", "Produto"], fallback_index=6)
        price_col = detect_column(raw, ["Preço venda", "Preco venda", "Preço Venda", "PRECO VENDA"], fallback_index=11)

        df = raw.copy()
        validated = df[sku_col].apply(validate_sku_format)
        df["SKU_NORMALIZADO"] = validated.apply(lambda x: x[0])
        df["SKU_SEVERIDADE"] = validated.apply(lambda x: x[1])
        df["SKU_MENSAGEM"] = validated.apply(lambda x: x[2])
        df["SKU_ORIGINAL"] = df[sku_col].astype(str)
        df["PRECO_TABELA_007"] = df[price_col].apply(parse_money)

        log = []
        for idx, row in df.iterrows():
            if row["SKU_SEVERIDADE"] != "OK" and (row["SKU_ORIGINAL"] or "").lower() not in {"nan", "none", ""}:
                log.append({
                    "Tipo": "FORMATO_SKU_TABELA_007",
                    "Severidade": row["SKU_SEVERIDADE"],
                    "SKU": row["SKU_NORMALIZADO"] or "",
                    "Mensagem": "Formato de SKU na tabela 007 exige revisão." if row["SKU_SEVERIDADE"] == "ATENÇÃO" else "Formato de SKU inválido na tabela 007.",
                    "Detalhe": f"Linha Excel aproximada: {idx + 2}; valor original: {row['SKU_ORIGINAL']}; {row['SKU_MENSAGEM']}",
                })
        dupes = df[df["SKU_NORMALIZADO"].notna() & df.duplicated("SKU_NORMALIZADO", keep=False)]
        if not dupes.empty:
            qtd = dupes["SKU_NORMALIZADO"].nunique()
            log.append({
                "Tipo": "DUPLICIDADE_TABELA_007",
                "Severidade": "ATENÇÃO",
                "SKU": "",
                "Mensagem": f"Há {qtd} SKU(s) duplicado(s) na tabela 007.",
                "Detalhe": "O app usará a primeira ocorrência para cálculo. Revise se houver preços diferentes.",
            })

        calc = df[df["SKU_NORMALIZADO"].notna()].drop_duplicates("SKU_NORMALIZADO", keep="first")
        return calc, raw, log, sheet


def prepare_lista_ml(lista_ml_file, config: ProcessConfig) -> Tuple[pd.DataFrame, list, str]:
    """Prepara Lista ML revisada.

    Usa leitura rápida por colunas fixas A:C. Se a estrutura mudar, cai no método
    flexível antigo com pandas para tentar localizar colunas pelo cabeçalho.
    """
    try:
        return _read_fast_lista_ml(lista_ml_file, config)
    except Exception:
        _safe_seek(lista_ml_file)
        raw, sheet = read_excel_flexible(lista_ml_file)
        sku_col = detect_column(raw, ["SKU", "Cod.Produto", "Cod Produto"], fallback_index=0)
        ml_col = detect_column(raw, ["Preço ML", "Preco ML", "PREÇO ML", "ML"], fallback_index=1)
        cost_col = detect_column(raw, ["Custo Médio", "Custo Medio", "CUSTO MEDIO", "Custo"], fallback_index=2)

        df = raw.copy()
        validated = df[sku_col].apply(validate_sku_format)
        df["SKU"] = validated.apply(lambda x: x[0])
        df["SKU_SEVERIDADE"] = validated.apply(lambda x: x[1])
        df["SKU_MENSAGEM"] = validated.apply(lambda x: x[2])
        df["SKU_ORIGINAL"] = df[sku_col].astype(str)
        df["PRECO_ML"] = df[ml_col].apply(parse_money)
        df["CUSTO_MEDIO"] = df[cost_col].apply(parse_money)

        log = []
        for idx, row in df.iterrows():
            if row["SKU_SEVERIDADE"] != "OK":
                log.append({
                    "Tipo": "FORMATO_SKU_LISTA_ML",
                    "Severidade": row["SKU_SEVERIDADE"],
                    "SKU": row["SKU"] or "",
                    "Mensagem": "Formato de SKU na Lista ML revisada exige revisão." if row["SKU_SEVERIDADE"] == "ATENÇÃO" else "Formato de SKU inválido na Lista ML revisada.",
                    "Detalhe": f"Linha Excel aproximada: {idx + 2}; valor original: {row['SKU_ORIGINAL']}; {row['SKU_MENSAGEM']}",
                })

        df = df[df["SKU"].notna()].copy()
        dupes = df[df.duplicated("SKU", keep=False)]
        if not dupes.empty:
            for sku, grp in dupes.groupby("SKU"):
                prices = sorted(set([round(x, 2) for x in grp["PRECO_ML"].dropna().tolist()]))
                costs = sorted(set([round(x, 2) for x in grp["CUSTO_MEDIO"].dropna().tolist()]))
                log.append({
                    "Tipo": "DUPLICIDADE_LISTA_ML",
                    "Severidade": "ATENÇÃO",
                    "SKU": sku,
                    "Mensagem": "SKU duplicado na Lista ML revisada.",
                    "Detalhe": f"Preços ML encontrados: {prices}; custos encontrados: {costs}. Usada a {config.usar_ocorrencia_duplicada} ocorrência.",
                })
            keep = "last" if config.usar_ocorrencia_duplicada == "ultima" else "first"
            df = df.drop_duplicates("SKU", keep=keep)

        return df[["SKU", "PRECO_ML", "CUSTO_MEDIO"]], log, sheet

# -----------------------------
# Processamento principal
# -----------------------------

def processar_precos(tabela_007_file, lista_ml_file, relatorio_anterior_file=None, config: Optional[ProcessConfig] = None, data_alteracao: Optional[date] = None) -> ProcessResult:
    config = config or ProcessConfig()
    data_alteracao = data_alteracao or date.today()
    data_str = data_alteracao.strftime("%d/%m/%Y")

    tabela, tabela_raw, log_tabela, sheet_tabela = prepare_tabela_007(tabela_007_file)
    lista_ml, log_ml, sheet_ml = prepare_lista_ml(lista_ml_file, config)
    log_rows = log_tabela + log_ml

    merged = lista_ml.merge(
        tabela,
        left_on="SKU",
        right_on="SKU_NORMALIZADO",
        how="left",
        suffixes=("_ML", "_T007"),
    )

    records = []
    matched_table_rows = []

    for _, row in merged.iterrows():
        sku = row["SKU"]
        preco_tabela = row.get("PRECO_TABELA_007")
        preco_ml = row.get("PRECO_ML")
        custo = row.get("CUSTO_MEDIO")

        obs = []
        acao = ""
        preco_final = None

        preco_tabela_sem_cadastro = (
            preco_tabela is not None
            and not pd.isna(preco_tabela)
            and abs(float(preco_tabela) - float(config.preco_007_sem_cadastro)) <= config.tolerancia
        )

        if pd.isna(preco_tabela):
            log_rows.append({
                "Tipo": "SKU_NAO_ENCONTRADO_TABELA_007",
                "Severidade": "CRÍTICO",
                "SKU": sku,
                "Mensagem": "SKU da Lista ML não encontrado na tabela 007 ou sem preço comercial válido.",
                "Detalhe": "Não será enviado ao Protheus até cadastro/preço ser validado.",
            })
            obs.append("SKU não encontrado na tabela 007 ou preço comercial inválido")
            acao = "NÃO PROCESSADO"
        elif preco_ml is None or pd.isna(preco_ml) or preco_ml <= 0:
            if preco_tabela_sem_cadastro:
                acao = "NÃO PROCESSADO - 007 SEM CADASTRO E ML INVÁLIDO"
                obs.append("Preço 007 igual a 0,01 indica ausência de cadastro; Preço ML inválido impede cálculo")
                log_rows.append({
                    "Tipo": "PRECO_007_SEM_CADASTRO_E_ML_INVALIDO",
                    "Severidade": "CRÍTICO",
                    "SKU": sku,
                    "Mensagem": "Preço 007 igual a 0,01 e preço ML inválido.",
                    "Detalhe": f"Preço 007 informado: {preco_tabela}; Preço ML informado: {preco_ml}. Item não enviado ao Protheus até revisão.",
                })
            else:
                preco_final = float(preco_tabela)
                acao = "MANTER - ML INVÁLIDO"
                obs.append("Preço ML inválido/zerado/negativo; mantido preço da tabela 007")
                log_rows.append({
                    "Tipo": "PRECO_ML_INVALIDO",
                    "Severidade": "CRÍTICO",
                    "SKU": sku,
                    "Mensagem": "Preço ML inválido, zerado ou negativo.",
                    "Detalhe": f"Preço ML informado: {preco_ml}. Preço 007 mantido: {preco_tabela}.",
                })
        elif preco_tabela_sem_cadastro:
            preco_final = round(float(preco_ml) * (1 - config.desconto_ml), 2)
            acao = "ALTERAR - 007 SEM CADASTRO (ML -10%)"
            obs.append("Preço 007 igual a 0,01 tratado como ausência de cadastro; aplicado Preço ML x 0,9")
            log_rows.append({
                "Tipo": "PRECO_007_SEM_CADASTRO",
                "Severidade": "ATENÇÃO",
                "SKU": sku,
                "Mensagem": "Preço 007 igual a 0,01 tratado como ausência de preço de venda cadastrado.",
                "Detalhe": f"Preço 007 atual: {preco_tabela}; Preço ML: {preco_ml}; novo preço 007 calculado: {preco_final}.",
            })
        elif float(preco_tabela) > float(preco_ml):
            preco_final = round(float(preco_ml) * (1 - config.desconto_ml), 2)
            acao = "ALTERAR - 10% ABAIXO ML"
        else:
            preco_final = round(float(preco_tabela), 2)
            acao = "MANTER - 007 <= ML"

        custo_valido = custo is not None and not pd.isna(custo) and custo > 0
        lucro = margem = markup = roi = None
        prejuizo = False
        faixa_margem = "CUSTO INVÁLIDO/AUSENTE"
        status_alerta = "OK"

        if preco_final is not None and custo_valido:
            # Regra financeira oficial: lucro bruto = preço de venda final da tabela 007 - custo médio.
            lucro = round(float(preco_final) - float(custo), 2)
            markup = round(float(preco_final) / float(custo), 4)
            margem = round(lucro / float(preco_final), 4) if preco_final else None
            roi = round(lucro / float(custo), 4) if custo else None
            prejuizo = lucro < 0

            if prejuizo:
                status_alerta = "CRÍTICO - PREJUÍZO"
                faixa_margem = "PREJUÍZO"
            elif markup < config.markup_minimo:
                status_alerta = "CRÍTICO - MARKUP < 120%"
            elif margem is not None and margem < config.margem_critica:
                status_alerta = "CRÍTICO - MARGEM < 35%"
            elif margem is not None and config.margem_critica <= margem <= config.margem_atencao:
                status_alerta = "ATENÇÃO - MARGEM 35% A 50%"
            else:
                status_alerta = "OK"

            if faixa_margem not in {"PREJUÍZO"}:
                if margem is None:
                    faixa_margem = "N/A"
                elif margem < config.margem_critica:
                    faixa_margem = "<35%"
                elif margem <= config.margem_atencao:
                    faixa_margem = "35% a 50%"
                else:
                    faixa_margem = ">50%"
        elif preco_final is not None:
            status_alerta = "ATENÇÃO - CUSTO INVÁLIDO/AUSENTE"
            obs.append("Custo médio inválido, zerado ou ausente; métricas financeiras não calculadas")
            log_rows.append({
                "Tipo": "CUSTO_MEDIO_INVALIDO",
                "Severidade": "ATENÇÃO",
                "SKU": sku,
                "Mensagem": "Custo médio inválido, zerado ou ausente.",
                "Detalhe": f"Custo informado: {custo}.",
            })

        records.append({
            "SKU": sku,
            "Preço Tabela 007 Atual": round(float(preco_tabela), 2) if preco_tabela is not None and not pd.isna(preco_tabela) else None,
            "Preço ML": round(float(preco_ml), 2) if preco_ml is not None and not pd.isna(preco_ml) else None,
            "Custo Médio": round(float(custo), 2) if custo is not None and not pd.isna(custo) else None,
            "Preço Final Sugerido 007": preco_final,
            "Ação": acao,
            "Markup": markup,
            "Lucro Bruto": lucro,
            "Margem de Lucro": margem,
            "ROI": roi,
            "Prejuízo?": "SIM" if prejuizo else "NÃO",
            "Faixa Margem": faixa_margem,
            "Status Alerta": status_alerta,
            "Data Alteração": data_str,
            "Observação": "; ".join(obs),
        })

        if pd.notna(row.get("SKU_NORMALIZADO")) and preco_final is not None:
            # Guarda a linha original da tabela 007, mas com preço venda atualizado se a coluna existir.
            matched_table_rows.append(row.to_dict())

    analise = pd.DataFrame(records)

    alertas = analise[
        analise["Status Alerta"].ne("OK") |
        analise["Prejuízo?"].eq("SIM") |
        analise["Observação"].astype(str).ne("")
    ].copy()

    protheus = analise[
        analise["Preço Final Sugerido 007"].notna() & analise["SKU"].notna()
    ][["SKU", "Preço Final Sugerido 007", "Data Alteração"]].copy()
    protheus.columns = ["SKU", "Preço", "Data"]
    protheus["SKU"] = protheus["SKU"].astype(str).apply(normalize_sku)
    protheus["Preço"] = protheus["Preço"].apply(lambda x: money_dot(x))

    protheus_013 = analise[
        analise["Preço Final Sugerido 007"].notna() & analise["SKU"].notna()
    ][["SKU", "Preço Final Sugerido 007", "Data Alteração"]].copy()
    protheus_013.columns = ["SKU", "Preço", "Data"]
    protheus_013["SKU"] = protheus_013["SKU"].astype(str).apply(normalize_sku)
    protheus_013["Preço"] = protheus_013["Preço"].apply(lambda x: money_dot(round(float(x) * (1 + config.acrescimo_tabela_013), 2)))

    comparativo = gerar_comparativo_anterior(analise, relatorio_anterior_file, config) if relatorio_anterior_file else pd.DataFrame()

    resumo_dict = {
        "Data de geração": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "Aba tabela 007 lida": sheet_tabela,
        "Aba Lista ML lida": sheet_ml,
        "Código tabela Protheus 007": str(config.tabela_codigo).zfill(3),
        "Código tabela Protheus 013": str(config.tabela_codigo_013).zfill(3),
        "Acréscimo aplicado na tabela 013": f"{config.acrescimo_tabela_013:.2%}",
        "SKUs na Lista ML revisada": int(len(lista_ml)),
        "SKUs encontrados na tabela 007": int(analise["Preço Tabela 007 Atual"].notna().sum()),
        "SKUs não encontrados ou sem preço 007": int(analise["Preço Tabela 007 Atual"].isna().sum()),
        "Preços alterados": int(analise["Ação"].astype(str).str.startswith("ALTERAR").sum()),
        "Preços calculados por 007 sem cadastro (0,01)": int((analise["Ação"] == "ALTERAR - 007 SEM CADASTRO (ML -10%)").sum()),
        "Preços mantidos 007 <= ML": int((analise["Ação"] == "MANTER - 007 <= ML").sum()),
        "Preços mantidos por ML inválido": int((analise["Ação"] == "MANTER - ML INVÁLIDO").sum()),
        "Itens com prejuízo vs custo médio": int((analise["Prejuízo?"] == "SIM").sum()),
        "Itens com markup menor que 120%": int((analise["Markup"].notna() & (analise["Markup"] < config.markup_minimo)).sum()),
        "Itens com margem menor que 35%": int((analise["Margem de Lucro"].notna() & (analise["Margem de Lucro"] < config.margem_critica)).sum()),
        "Itens com margem entre 35% e 50%": int((analise["Margem de Lucro"].notna() & (analise["Margem de Lucro"] >= config.margem_critica) & (analise["Margem de Lucro"] <= config.margem_atencao)).sum()),
        "Itens com custo médio inválido/ausente": int((analise["Custo Médio"].isna() | (analise["Custo Médio"] <= 0)).sum()),
    }
    if not comparativo.empty:
        resumo_dict["Novos SKUs vs relatório anterior"] = int((comparativo["Tipo Alteração"] == "NOVO NA LISTA ML").sum())
        resumo_dict["SKUs removidos vs relatório anterior"] = int((comparativo["Tipo Alteração"] == "SAIU DA LISTA ML").sum())
        resumo_dict["SKUs com preço ML alterado"] = int((comparativo["Tipo Alteração"] == "ALTEROU PREÇO ML").sum())
        resumo_dict["SKUs com preço final alterado"] = int((comparativo["Tipo Alteração"] == "ALTEROU PREÇO FINAL").sum())

    resumo = pd.DataFrame([{"Indicador": k, "Valor": v} for k, v in resumo_dict.items()])
    log = pd.DataFrame(log_rows, columns=["Tipo", "Severidade", "SKU", "Mensagem", "Detalhe"])

    # Tabela filtrada em estrutura analítica; a aba operacional do relatório final é a análise + Protheus.
    matched_table = analise.copy()
    log_text = build_log_txt_from_frames(resumo, log, analise, protheus, protheus_013)

    return ProcessResult(
        analise=analise,
        alertas=alertas,
        log=log,
        resumo=resumo,
        comparativo=comparativo,
        protheus=protheus,
        protheus_013=protheus_013,
        matched_table=matched_table,
        resumo_dict=resumo_dict,
        log_text=log_text,
    )


def gerar_comparativo_anterior(analise_atual: pd.DataFrame, relatorio_anterior_file, config: ProcessConfig) -> pd.DataFrame:
    try:
        prev, _ = read_excel_flexible(relatorio_anterior_file, preferred_sheet="ANALISE_PRECOS")
    except Exception:
        return pd.DataFrame([{
            "Tipo Alteração": "ERRO_COMPARATIVO",
            "SKU": "",
            "Campo": "",
            "Valor Anterior": "",
            "Valor Atual": "",
            "Observação": "Não foi possível ler a aba ANALISE_PRECOS do relatório anterior.",
        }])

    try:
        sku_col = detect_column(prev, ["SKU"], fallback_index=0)
        prev = prev.copy()
        prev["SKU"] = prev[sku_col].apply(normalize_sku)
        prev = prev[prev["SKU"].notna()].drop_duplicates("SKU", keep="last")

        def find_optional(cands):
            try:
                return detect_column(prev, cands)
            except Exception:
                return None

        col_ml = find_optional(["Preço ML", "Preco ML", "PRECO_ML"])
        col_custo = find_optional(["Custo Médio", "Custo Medio", "CUSTO_MEDIO"])
        col_final = find_optional(["Preço Final Sugerido 007", "Preço Final", "Novo Preço", "PRECO_FINAL"])

        prev_slim = pd.DataFrame({"SKU": prev["SKU"]})
        if col_ml: prev_slim["Preço ML Anterior"] = prev[col_ml].apply(parse_money)
        if col_custo: prev_slim["Custo Médio Anterior"] = prev[col_custo].apply(parse_money)
        if col_final: prev_slim["Preço Final Anterior"] = prev[col_final].apply(parse_money)

        atual = analise_atual[["SKU", "Preço ML", "Custo Médio", "Preço Final Sugerido 007"]].copy()
        atual = atual.drop_duplicates("SKU", keep="last")

        cmp = atual.merge(prev_slim, on="SKU", how="outer", indicator=True)
        rows = []
        for _, r in cmp.iterrows():
            sku = r["SKU"]
            if r["_merge"] == "left_only":
                rows.append({"Tipo Alteração": "NOVO NA LISTA ML", "SKU": sku, "Campo": "SKU", "Valor Anterior": "", "Valor Atual": sku, "Observação": "SKU não constava no relatório anterior."})
            elif r["_merge"] == "right_only":
                rows.append({"Tipo Alteração": "SAIU DA LISTA ML", "SKU": sku, "Campo": "SKU", "Valor Anterior": sku, "Valor Atual": "", "Observação": "SKU constava antes e não está na lista atual."})
            else:
                comparisons = [
                    ("ALTEROU PREÇO ML", "Preço ML", r.get("Preço ML Anterior"), r.get("Preço ML")),
                    ("ALTEROU CUSTO MÉDIO", "Custo Médio", r.get("Custo Médio Anterior"), r.get("Custo Médio")),
                    ("ALTEROU PREÇO FINAL", "Preço Final Sugerido 007", r.get("Preço Final Anterior"), r.get("Preço Final Sugerido 007")),
                ]
                for tipo, campo, old, new in comparisons:
                    if pd.isna(old) and pd.isna(new):
                        continue
                    if pd.isna(old) != pd.isna(new) or abs(float(old or 0) - float(new or 0)) > config.tolerancia:
                        rows.append({"Tipo Alteração": tipo, "SKU": sku, "Campo": campo, "Valor Anterior": old, "Valor Atual": new, "Observação": "Alteração identificada contra o relatório anterior."})
        return pd.DataFrame(rows)
    except Exception as exc:
        return pd.DataFrame([{
            "Tipo Alteração": "ERRO_COMPARATIVO",
            "SKU": "",
            "Campo": "",
            "Valor Anterior": "",
            "Valor Atual": "",
            "Observação": f"Erro ao comparar relatório anterior: {exc}",
        }])




def build_log_txt_from_frames(resumo: pd.DataFrame, log: pd.DataFrame, analise: pd.DataFrame, protheus_007: pd.DataFrame, protheus_013: pd.DataFrame) -> str:
    linhas = []
    linhas.append("LOG DE PROCESSAMENTO - AJUSTE TABELAS 007/013 x MERCADO LIVRE")
    linhas.append(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    linhas.append("")
    linhas.append("RESUMO")
    linhas.append("------")
    if resumo is not None and not resumo.empty:
        for _, r in resumo.iterrows():
            linhas.append(f"- {r.get('Indicador')}: {r.get('Valor')}")
    linhas.append(f"- Linhas geradas no arquivo Protheus 007: {len(protheus_007)}")
    linhas.append(f"- Linhas geradas no arquivo Protheus 013: {len(protheus_013)}")
    linhas.append("")

    # Validação defensiva: Protheus precisa sair com SKUs de 5 dígitos.
    sku_invalidos_saida = []
    for nome, df in [("007", protheus_007), ("013", protheus_013)]:
        if df is None or df.empty:
            continue
        for sku in df["SKU"].astype(str).tolist():
            if not re.fullmatch(r"\d{5}", sku):
                sku_invalidos_saida.append((nome, sku))
    if sku_invalidos_saida:
        linhas.append("ALERTA CRÍTICO - FORMATO DE SKU NA SAÍDA PROTHEUS")
        linhas.append("---------------------------------------------------")
        for tabela, sku in sku_invalidos_saida[:200]:
            linhas.append(f"- Tabela {tabela}: SKU fora do padrão de 5 dígitos -> {sku}")
        if len(sku_invalidos_saida) > 200:
            linhas.append(f"- ... mais {len(sku_invalidos_saida) - 200} ocorrência(s).")
        linhas.append("")

    linhas.append("INTERCORRÊNCIAS")
    linhas.append("---------------")
    if log is None or log.empty:
        linhas.append("Nenhuma intercorrência registrada no processamento.")
    else:
        for _, r in log.iterrows():
            tipo = r.get("Tipo", "")
            sev = r.get("Severidade", "")
            sku = r.get("SKU", "")
            msg = r.get("Mensagem", "")
            det = r.get("Detalhe", "")
            linhas.append(f"[{sev}] {tipo} | SKU: {sku} | {msg}")
            if det:
                linhas.append(f"    Detalhe: {det}")
    linhas.append("")

    criticos = 0 if log is None or log.empty else int((log["Severidade"].astype(str).str.upper() == "CRÍTICO").sum())
    atencao = 0 if log is None or log.empty else int((log["Severidade"].astype(str).str.upper() == "ATENÇÃO").sum())
    prejuizos = 0 if analise is None or analise.empty else int((analise["Prejuízo?"] == "SIM").sum())
    linhas.append("SÍNTESE DE RISCO")
    linhas.append("----------------")
    linhas.append(f"- Intercorrências críticas: {criticos}")
    linhas.append(f"- Intercorrências de atenção: {atencao}")
    linhas.append(f"- Itens com prejuízo vs custo médio: {prejuizos}")
    linhas.append("")
    linhas.append("Observação: o arquivo TXT é apenas um log de auditoria. A análise completa permanece no relatório Excel.")
    return "\n".join(linhas)


def build_log_txt(result: ProcessResult) -> bytes:
    return result.log_text.encode("utf-8-sig")

# -----------------------------
# Escrita dos arquivos finais
# -----------------------------

def write_df(ws, df: pd.DataFrame, start_row=1, start_col=1, table_name: Optional[str] = None):
    # Header
    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(start_row, j, str(col))
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Body
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            val = row[col]
            if pd.isna(val):
                val = None
            cell = ws.cell(i, j, val)
            if "SKU" in str(col).upper():
                cell.number_format = "00000"
            elif str(col).lower() in {"markup", "margem de lucro", "roi"}:
                cell.number_format = "0.00%"
            elif "preço" in str(col).lower() or "custo" in str(col).lower() or "lucro" in str(col).lower():
                cell.number_format = '#,##0.00'

    max_row = start_row + len(df)
    max_col = start_col + len(df.columns) - 1
    if table_name and len(df) > 0:
        ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(max_col)}{max_row}"
        tab = Table(displayName=table_name[:250], ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)

    for col_idx in range(start_col, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 10
        for cell in ws[letter]:
            try:
                max_len = max(max_len, min(len(str(cell.value)) if cell.value is not None else 0, 45))
            except Exception:
                pass
        ws.column_dimensions[letter].width = max_len + 2

    ws.freeze_panes = ws.cell(start_row + 1, start_col)
    return ws


def apply_report_formatting(wb: Workbook):
    """Formatação leve do relatório.

    Evita percorrer todas as células aplicando bordas, que era o principal custo
    adicional da versão com formatação mais pesada. As tabelas do Excel já recebem
    estilo próprio em write_df; aqui mantemos gridlines e destaques de alerta.
    """
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = True

    if "ALERTAS_FINANCEIROS" in wb.sheetnames:
        ws = wb["ALERTAS_FINANCEIROS"]
        header = [c.value for c in ws[1]]
        if "Status Alerta" in header:
            col = header.index("Status Alerta") + 1
            for r in range(2, ws.max_row + 1):
                val = str(ws.cell(r, col).value or "")
                if "CRÍTICO" in val:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = PatternFill("solid", fgColor="F8CBAD")
                elif "ATENÇÃO" in val:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = PatternFill("solid", fgColor="FFF2CC")

def build_report_excel(result: ProcessResult) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("RESUMO", result.resumo, "ResumoTable"),
        ("ANALISE_PRECOS", result.analise, "AnalisePrecosTable"),
        ("ALERTAS_FINANCEIROS", result.alertas, "AlertasFinanceirosTable"),
        ("COMPARATIVO_ANTERIOR", result.comparativo, "ComparativoAnteriorTable"),
        ("LOG_AJUSTE_ML_007", result.log, "LogAjusteTable"),
        ("SAIDA_PROTHEUS_007", result.protheus, "SaidaProtheus007Table"),
        ("SAIDA_PROTHEUS_013", result.protheus_013, "SaidaProtheus013Table"),
    ]

    for sheet_name, df, table_name in sheets:
        ws = wb.create_sheet(sheet_name)
        if df is not None and not df.empty:
            write_df(ws, df, table_name=table_name)
        else:
            ws["A1"] = "Sem registros."
            ws["A1"].font = Font(bold=True)

    apply_report_formatting(wb)

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_protheus_excel(result: ProcessResult, config: Optional[ProcessConfig] = None, tabela: str = "007") -> bytes:
    config = config or ProcessConfig()
    tabela = str(tabela).zfill(3)

    if tabela == str(config.tabela_codigo_013).zfill(3) or tabela == "013":
        df_saida = result.protheus_013
    else:
        df_saida = result.protheus

    wb = Workbook()
    ws = wb.active
    ws.title = tabela

    # A1 e os SKUs precisam ser valores numéricos com formato personalizado,
    # igual ao modelo usado pelos analistas: 000 para tabela e 00000 para SKU.
    # Assim o Excel exibe zeros à esquerda sem transformar o campo em texto.
    ws["A1"] = int(tabela)
    ws["A1"].number_format = "000"
    ws["A1"].font = Font(bold=True)

    for out_idx, (_, row) in enumerate(df_saida.iterrows(), start=2):
        sku_norm = normalize_sku(row["SKU"])
        sku_cell = ws.cell(out_idx, 1, int(sku_norm) if sku_norm and sku_norm.isdigit() else sku_norm)
        sku_cell.number_format = "00000"
        ws.cell(out_idx, 2, str(row["Preço"])).number_format = "@"  # ponto decimal como texto para preservar 25.99
        ws.cell(out_idx, 3, str(row["Data"])).number_format = "@"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.sheet_view.showGridLines = True

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()




def _csv_value(value) -> str:
    """Escapa valores para CSV simples separado por ponto e vírgula."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value)
    if any(ch in text for ch in [';', '"', '\n', '\r']):
        text = '"' + text.replace('"', '""') + '"'
    return text


def normalize_table_code(value) -> str:
    """Normaliza código de tabela Protheus para 3 dígitos em CSV."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "000"
    text = str(value).strip()
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    digits = re.sub(r"\D", "", text)
    if digits:
        return digits.zfill(3)[-3:]
    return text.zfill(3)[:3]


def build_protheus_csv_from_df(df_saida: pd.DataFrame, tabela: str) -> bytes:
    """Gera CSV de subida Protheus no layout aceito pelo importador DA1.

    Formato obrigatório validado contra o modelo manual que funcionou:
    - Linha 1: 007;;;  -> 4 campos
    - Linhas seguintes: SKU;PRECO;DATA;  -> 4 campos, com ; final

    Não use csv.writer aqui: alguns leitores/importadores removem o campo vazio
    final dependendo da configuração. Montamos a linha de forma explícita para
    garantir que o último caractere útil de cada linha seja sempre ';'.
    """
    tabela_norm = normalize_table_code(tabela)
    linhas = [f"{tabela_norm};;;"]

    if df_saida is not None and not df_saida.empty:
        for _, row in df_saida.iterrows():
            sku_norm = normalize_sku(row.get("SKU"))
            preco = row.get("Preço")
            data = row.get("Data")

            sku_txt = _csv_value(sku_norm)
            preco_txt = _csv_value(preco)
            data_txt = _csv_value(data)

            # O ; final é obrigatório para o Protheus criar a 4ª posição do array.
            linhas.append(f"{sku_txt};{preco_txt};{data_txt};")

    conteudo = "\r\n".join(linhas) + "\r\n"
    return conteudo.encode("utf-8")

def build_protheus_csv(result: ProcessResult, config: Optional[ProcessConfig] = None, tabela: str = "007") -> bytes:
    config = config or ProcessConfig()
    tabela = str(tabela).zfill(3)
    if tabela == str(config.tabela_codigo_013).zfill(3) or tabela == "013":
        df_saida = result.protheus_013
    else:
        df_saida = result.protheus
    return build_protheus_csv_from_df(df_saida, tabela)


def build_zip(report_bytes: bytes, protheus_007_bytes: bytes, protheus_013_bytes: bytes, log_txt_bytes: bytes, report_name: str, protheus_007_name: str, protheus_013_name: str, log_name: str) -> bytes:
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(report_name, report_bytes)
        zf.writestr(protheus_007_name, protheus_007_bytes)
        zf.writestr(protheus_013_name, protheus_013_bytes)
        zf.writestr(log_name, log_txt_bytes)
    return bio.getvalue()



# ============================================================
# MÓDULO 2 - PRECIFICAÇÃO COMUM
# ============================================================

@dataclass
class CommonPricingConfig:
    tabela_001: str = "001"
    tabela_004: str = "004"
    tabela_012: str = "012"
    tabela_007: str = "007"
    tabela_013: str = "013"
    desconto_007_sobre_001: float = 0.16
    acrescimo_013_sobre_007: float = 0.12
    usar_ocorrencia_duplicada: str = "ultima"  # ultima | primeira


@dataclass
class CommonPricingResult:
    base: pd.DataFrame
    analise: pd.DataFrame
    log: pd.DataFrame
    resumo: pd.DataFrame
    saidas: Dict[str, pd.DataFrame]
    resumo_dict: Dict[str, int | float | str]
    log_text: str


def _read_fast_precificacao_comum(input_file, config: CommonPricingConfig) -> Tuple[pd.DataFrame, list, str]:
    """Lê planilha de Precificação Comum.

    Estrutura esperada: A=SKU, B=Custo Médio, C=Preço 001.
    A primeira linha pode conter cabeçalho.
    """
    _safe_seek(input_file)
    wb = load_workbook(_excel_source(input_file), read_only=True, data_only=True)
    sheet = wb.sheetnames[0]
    ws = wb[sheet]

    rows = []
    log = []
    for excel_row, values in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True), start=2):
        raw_sku, raw_custo, raw_preco_001 = values
        sku, sev, msg = validate_sku_format(raw_sku)
        if sku is None and (raw_sku is None or str(raw_sku).strip().lower() in {"", "nan", "none", "null"}):
            continue

        if sev != "OK":
            log.append({
                "Tipo": "FORMATO_SKU_PRECIFICACAO_COMUM",
                "Severidade": sev,
                "SKU": sku or "",
                "Mensagem": "Formato de SKU na planilha de precificação comum exige revisão." if sev == "ATENÇÃO" else "Formato de SKU inválido na planilha de precificação comum.",
                "Detalhe": f"Linha Excel aproximada: {excel_row}; valor original: {raw_sku}; {msg}",
            })
        if sku:
            rows.append({
                "SKU": sku,
                "CUSTO_MEDIO": parse_money(raw_custo),
                "PRECO_001": parse_money(raw_preco_001),
                "SKU_ORIGINAL": str(raw_sku),
                "LINHA_EXCEL": excel_row,
            })

    try:
        wb.close()
    except Exception:
        pass

    df = pd.DataFrame(rows, columns=["SKU", "CUSTO_MEDIO", "PRECO_001", "SKU_ORIGINAL", "LINHA_EXCEL"])
    if df.empty:
        return df, log, sheet

    dupes = df[df.duplicated("SKU", keep=False)]
    if not dupes.empty:
        for sku, grp in dupes.groupby("SKU"):
            custos = sorted(set([round(x, 2) for x in grp["CUSTO_MEDIO"].dropna().tolist()]))
            precos = sorted(set([round(x, 2) for x in grp["PRECO_001"].dropna().tolist()]))
            linhas = grp["LINHA_EXCEL"].tolist()
            log.append({
                "Tipo": "DUPLICIDADE_PRECIFICACAO_COMUM",
                "Severidade": "ATENÇÃO",
                "SKU": sku,
                "Mensagem": "SKU duplicado na planilha de Precificação Comum.",
                "Detalhe": f"Linhas: {linhas}; custos encontrados: {custos}; preços 001 encontrados: {precos}. Usada a {config.usar_ocorrencia_duplicada} ocorrência.",
            })
        keep = "last" if config.usar_ocorrencia_duplicada == "ultima" else "first"
        df = df.drop_duplicates("SKU", keep=keep)

    return df[["SKU", "CUSTO_MEDIO", "PRECO_001", "SKU_ORIGINAL", "LINHA_EXCEL"]], log, sheet


def prepare_precificacao_comum(input_file, config: CommonPricingConfig) -> Tuple[pd.DataFrame, list, str]:
    """Prepara base de Precificação Comum, com fallback flexível por cabeçalho."""
    try:
        return _read_fast_precificacao_comum(input_file, config)
    except Exception:
        _safe_seek(input_file)
        raw, sheet = read_excel_flexible(input_file)
        sku_col = detect_column(raw, ["SKU", "Cod.Produto", "Cod Produto"], fallback_index=0)
        custo_col = detect_column(raw, ["Custo Médio", "Custo Medio", "Custo", "CUSTO MEDIO"], fallback_index=1)
        preco_col = detect_column(raw, ["Preço 001", "Preco 001", "001", "Preço Tabela 001", "Preco Tabela 001"], fallback_index=2)

        df = raw.copy()
        validated = df[sku_col].apply(validate_sku_format)
        df["SKU"] = validated.apply(lambda x: x[0])
        df["SKU_SEVERIDADE"] = validated.apply(lambda x: x[1])
        df["SKU_MENSAGEM"] = validated.apply(lambda x: x[2])
        df["SKU_ORIGINAL"] = df[sku_col].astype(str)
        df["CUSTO_MEDIO"] = df[custo_col].apply(parse_money)
        df["PRECO_001"] = df[preco_col].apply(parse_money)
        df["LINHA_EXCEL"] = df.index + 2

        log = []
        for _, row in df.iterrows():
            if row["SKU_SEVERIDADE"] != "OK" and str(row["SKU_ORIGINAL"]).lower() not in {"nan", "none", ""}:
                log.append({
                    "Tipo": "FORMATO_SKU_PRECIFICACAO_COMUM",
                    "Severidade": row["SKU_SEVERIDADE"],
                    "SKU": row["SKU"] or "",
                    "Mensagem": "Formato de SKU na planilha de precificação comum exige revisão." if row["SKU_SEVERIDADE"] == "ATENÇÃO" else "Formato de SKU inválido na planilha de precificação comum.",
                    "Detalhe": f"Linha Excel aproximada: {row['LINHA_EXCEL']}; valor original: {row['SKU_ORIGINAL']}; {row['SKU_MENSAGEM']}",
                })

        df = df[df["SKU"].notna()].copy()
        dupes = df[df.duplicated("SKU", keep=False)]
        if not dupes.empty:
            for sku, grp in dupes.groupby("SKU"):
                custos = sorted(set([round(x, 2) for x in grp["CUSTO_MEDIO"].dropna().tolist()]))
                precos = sorted(set([round(x, 2) for x in grp["PRECO_001"].dropna().tolist()]))
                linhas = grp["LINHA_EXCEL"].tolist()
                log.append({
                    "Tipo": "DUPLICIDADE_PRECIFICACAO_COMUM",
                    "Severidade": "ATENÇÃO",
                    "SKU": sku,
                    "Mensagem": "SKU duplicado na planilha de Precificação Comum.",
                    "Detalhe": f"Linhas: {linhas}; custos encontrados: {custos}; preços 001 encontrados: {precos}. Usada a {config.usar_ocorrencia_duplicada} ocorrência.",
                })
            keep = "last" if config.usar_ocorrencia_duplicada == "ultima" else "first"
            df = df.drop_duplicates("SKU", keep=keep)

        return df[["SKU", "CUSTO_MEDIO", "PRECO_001", "SKU_ORIGINAL", "LINHA_EXCEL"]], log, sheet


def processar_precificacao_comum(input_file, config: Optional[CommonPricingConfig] = None, data_alteracao: Optional[date] = None) -> CommonPricingResult:
    config = config or CommonPricingConfig()
    data_alteracao = data_alteracao or date.today()
    data_str = data_alteracao.strftime("%d/%m/%Y")

    base, log_rows, sheet = prepare_precificacao_comum(input_file, config)
    records = []

    for _, row in base.iterrows():
        sku = row.get("SKU")
        custo = row.get("CUSTO_MEDIO")
        preco_001 = row.get("PRECO_001")
        linha = row.get("LINHA_EXCEL")
        obs = []

        custo_valido = custo is not None and not pd.isna(custo) and float(custo) > 0
        preco_001_valido = preco_001 is not None and not pd.isna(preco_001) and float(preco_001) > 0

        if not custo_valido:
            obs.append("Custo médio inválido/zerado/ausente")
            log_rows.append({
                "Tipo": "CUSTO_INVALIDO_PRECIFICACAO_COMUM",
                "Severidade": "CRÍTICO",
                "SKU": sku or "",
                "Mensagem": "Custo médio inválido, zerado ou ausente.",
                "Detalhe": f"Linha Excel aproximada: {linha}; custo informado: {custo}. As tabelas 004 e 012 não serão geradas para este SKU.",
            })

        if not preco_001_valido:
            obs.append("Preço 001 inválido/zerado/ausente")
            log_rows.append({
                "Tipo": "PRECO_001_INVALIDO_PRECIFICACAO_COMUM",
                "Severidade": "CRÍTICO",
                "SKU": sku or "",
                "Mensagem": "Preço 001 inválido, zerado ou ausente.",
                "Detalhe": f"Linha Excel aproximada: {linha}; preço 001 informado: {preco_001}. As tabelas 001, 007 e 013 não serão geradas para este SKU.",
            })

        preco_001_calc = round(float(preco_001), 2) if preco_001_valido else None
        preco_004_calc = round(float(custo) / 2, 2) if custo_valido else None
        preco_012_calc = round(float(custo) / 2, 2) if custo_valido else None
        preco_007_calc = round(float(preco_001_calc) * (1 - config.desconto_007_sobre_001), 2) if preco_001_calc is not None else None
        preco_013_calc = round(float(preco_007_calc) * (1 + config.acrescimo_013_sobre_007), 2) if preco_007_calc is not None else None

        records.append({
            "SKU": sku,
            "Custo Médio": round(float(custo), 2) if custo is not None and not pd.isna(custo) else None,
            "Preço 001 Informado": round(float(preco_001), 2) if preco_001 is not None and not pd.isna(preco_001) else None,
            "Preço Tabela 001": preco_001_calc,
            "Preço Tabela 004": preco_004_calc,
            "Preço Tabela 012": preco_012_calc,
            "Preço Tabela 007": preco_007_calc,
            "Preço Tabela 013": preco_013_calc,
            "Data Alteração": data_str,
            "Status": "OK" if not obs else "VERIFICAR",
            "Observação": "; ".join(obs),
        })

    analise = pd.DataFrame(records)

    def _saida(codigo: str, preco_col: str) -> pd.DataFrame:
        if analise.empty or preco_col not in analise.columns:
            return pd.DataFrame(columns=["SKU", "Preço", "Data"])
        df = analise[analise[preco_col].notna() & analise["SKU"].notna()][["SKU", preco_col, "Data Alteração"]].copy()
        df.columns = ["SKU", "Preço", "Data"]
        df["SKU"] = df["SKU"].astype(str).apply(normalize_sku)
        df["Preço"] = df["Preço"].apply(lambda x: money_dot(x))
        return df

    saidas = {
        str(config.tabela_001).zfill(3): _saida(str(config.tabela_001).zfill(3), "Preço Tabela 001"),
        str(config.tabela_004).zfill(3): _saida(str(config.tabela_004).zfill(3), "Preço Tabela 004"),
        str(config.tabela_012).zfill(3): _saida(str(config.tabela_012).zfill(3), "Preço Tabela 012"),
        str(config.tabela_007).zfill(3): _saida(str(config.tabela_007).zfill(3), "Preço Tabela 007"),
        str(config.tabela_013).zfill(3): _saida(str(config.tabela_013).zfill(3), "Preço Tabela 013"),
    }

    resumo_dict = {
        "Data de geração": datetime.now().strftime("%d/%m/%Y %H:%M"),
        "Aba de origem lida": sheet,
        "SKUs válidos na base": int(len(base)),
        "SKUs com custo médio válido": int((analise["Custo Médio"].notna() & (analise["Custo Médio"] > 0)).sum()) if not analise.empty else 0,
        "SKUs com preço 001 válido": int((analise["Preço 001 Informado"].notna() & (analise["Preço 001 Informado"] > 0)).sum()) if not analise.empty else 0,
        "Linhas geradas tabela 001": int(len(saidas[str(config.tabela_001).zfill(3)])),
        "Linhas geradas tabela 004": int(len(saidas[str(config.tabela_004).zfill(3)])),
        "Linhas geradas tabela 012": int(len(saidas[str(config.tabela_012).zfill(3)])),
        "Linhas geradas tabela 007": int(len(saidas[str(config.tabela_007).zfill(3)])),
        "Linhas geradas tabela 013": int(len(saidas[str(config.tabela_013).zfill(3)])),
        "Desconto aplicado na 007 sobre a 001": f"{config.desconto_007_sobre_001:.2%}",
        "Acréscimo aplicado na 013 sobre a 007": f"{config.acrescimo_013_sobre_007:.2%}",
        "Intercorrências críticas": int(sum(1 for x in log_rows if str(x.get("Severidade", "")).upper() == "CRÍTICO")),
        "Intercorrências de atenção": int(sum(1 for x in log_rows if str(x.get("Severidade", "")).upper() == "ATENÇÃO")),
    }
    resumo = pd.DataFrame([{"Indicador": k, "Valor": v} for k, v in resumo_dict.items()])
    log = pd.DataFrame(log_rows, columns=["Tipo", "Severidade", "SKU", "Mensagem", "Detalhe"])
    log_text = build_common_log_txt_from_frames(resumo, log, saidas)

    return CommonPricingResult(
        base=base,
        analise=analise,
        log=log,
        resumo=resumo,
        saidas=saidas,
        resumo_dict=resumo_dict,
        log_text=log_text,
    )


def build_common_log_txt_from_frames(resumo: pd.DataFrame, log: pd.DataFrame, saidas: Dict[str, pd.DataFrame]) -> str:
    linhas = []
    linhas.append("LOG DE PROCESSAMENTO - PRECIFICAÇÃO COMUM PROTHEUS")
    linhas.append(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    linhas.append("")
    linhas.append("RESUMO")
    linhas.append("------")
    if resumo is not None and not resumo.empty:
        for _, r in resumo.iterrows():
            linhas.append(f"- {r.get('Indicador')}: {r.get('Valor')}")
    linhas.append("")

    sku_invalidos_saida = []
    for tabela, df in saidas.items():
        if df is None or df.empty:
            continue
        for sku in df["SKU"].astype(str).tolist():
            if not re.fullmatch(r"\d{5}", sku):
                sku_invalidos_saida.append((tabela, sku))
    if sku_invalidos_saida:
        linhas.append("ALERTA CRÍTICO - FORMATO DE SKU NA SAÍDA PROTHEUS")
        linhas.append("---------------------------------------------------")
        for tabela, sku in sku_invalidos_saida[:200]:
            linhas.append(f"- Tabela {tabela}: SKU fora do padrão de 5 dígitos -> {sku}")
        if len(sku_invalidos_saida) > 200:
            linhas.append(f"- ... mais {len(sku_invalidos_saida) - 200} ocorrência(s).")
        linhas.append("")

    linhas.append("INTERCORRÊNCIAS")
    linhas.append("---------------")
    if log is None or log.empty:
        linhas.append("Nenhuma intercorrência registrada no processamento.")
    else:
        for _, r in log.iterrows():
            tipo = r.get("Tipo", "")
            sev = r.get("Severidade", "")
            sku = r.get("SKU", "")
            msg = r.get("Mensagem", "")
            det = r.get("Detalhe", "")
            linhas.append(f"[{sev}] {tipo} | SKU: {sku} | {msg}")
            if det:
                linhas.append(f"    Detalhe: {det}")
    linhas.append("")
    linhas.append("Observação: o TXT é o log rápido. O relatório Excel contém a auditoria analítica e as saídas por tabela.")
    return "\n".join(linhas)


def build_common_log_txt(result: CommonPricingResult) -> bytes:
    return result.log_text.encode("utf-8-sig")


def build_protheus_excel_from_df(df_saida: pd.DataFrame, tabela: str) -> bytes:
    tabela = str(tabela).zfill(3)
    wb = Workbook()
    ws = wb.active
    ws.title = tabela

    ws["A1"] = int(tabela)
    ws["A1"].number_format = "000"
    ws["A1"].font = Font(bold=True)

    if df_saida is not None and not df_saida.empty:
        for out_idx, (_, row) in enumerate(df_saida.iterrows(), start=2):
            sku_norm = normalize_sku(row["SKU"])
            sku_cell = ws.cell(out_idx, 1, int(sku_norm) if sku_norm and str(sku_norm).isdigit() else sku_norm)
            sku_cell.number_format = "00000"
            ws.cell(out_idx, 2, str(row["Preço"])).number_format = "@"
            ws.cell(out_idx, 3, str(row["Data"])).number_format = "@"

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.sheet_view.showGridLines = True

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_common_protheus_excel(result: CommonPricingResult, tabela: str) -> bytes:
    tabela = str(tabela).zfill(3)
    return build_protheus_excel_from_df(result.saidas.get(tabela, pd.DataFrame(columns=["SKU", "Preço", "Data"])), tabela)




def build_common_protheus_csv(result: CommonPricingResult, tabela: str) -> bytes:
    tabela = str(tabela).zfill(3)
    return build_protheus_csv_from_df(result.saidas.get(tabela, pd.DataFrame(columns=["SKU", "Preço", "Data"])), tabela)


def build_common_report_excel(result: CommonPricingResult) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    sheets = [
        ("RESUMO", result.resumo, "ResumoComumTable"),
        ("ANALISE_PRECIFICACAO", result.analise, "AnalisePrecificacaoComumTable"),
        ("LOG_PRECIFICACAO", result.log, "LogPrecificacaoComumTable"),
    ]
    for tabela, df in result.saidas.items():
        sheets.append((f"SAIDA_PROTHEUS_{tabela}", df, f"SaidaProtheus{tabela}Table"))

    for sheet_name, df, table_name in sheets:
        ws = wb.create_sheet(sheet_name[:31])
        if df is not None and not df.empty:
            write_df(ws, df, table_name=table_name)
        else:
            ws["A1"] = "Sem registros."
            ws["A1"].font = Font(bold=True)
        ws.sheet_view.showGridLines = True

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def build_zip_generic(files: Dict[str, bytes]) -> bytes:
    bio = BytesIO()
    with zipfile.ZipFile(bio, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for name, content in files.items():
            zf.writestr(name, content)
    return bio.getvalue()


def build_common_zip(result: CommonPricingResult, data_nome: str) -> Tuple[bytes, Dict[str, bytes], Dict[str, str]]:
    report_name = f"relatorio_precificacao_comum_{data_nome}.xlsx"
    log_name = f"log_precificacao_comum_{data_nome}.txt"

    files: Dict[str, bytes] = {}
    names: Dict[str, str] = {}

    files[report_name] = build_common_report_excel(result)
    names["relatorio"] = report_name

    for tabela in result.saidas.keys():
        fname = f"protheus_tabela_{tabela}_{data_nome}.csv"
        files[fname] = build_common_protheus_csv(result, tabela)
        names[f"protheus_{tabela}"] = fname

    files[log_name] = build_common_log_txt(result)
    names["log"] = log_name

    zip_name = f"pacote_precificacao_comum_{data_nome}.zip"
    names["zip"] = zip_name
    zip_bytes = build_zip_generic(files)
    return zip_bytes, files, names
